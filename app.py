# app.py
from flask import Flask, render_template, jsonify, redirect, request, send_from_directory
import yfinance as yf
import pandas as pd
import threading
import requests
from io import StringIO
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import time as time_module
import time
import random

app = Flask(__name__)

# Configuration
LOG_FILE = "target_hit_log.xlsx"
DATA_FILE = "dailystock.xlsx"
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1qPeDQOzgiCrfp1h32KUyn5CHD509yR8E_ggxfjFtJOc/edit?usp=sharing"
USE_GOOGLE_SHEETS = True
print(f"USE_GOOGLE_SHEETS: {USE_GOOGLE_SHEETS}")
if USE_GOOGLE_SHEETS:
    print(f"GOOGLE_SHEET_URL: {GOOGLE_SHEET_URL}")

# In-memory storage
watchlists = {}
target_hit_logged = {}
monitoring_active = True

# Configure yfinance session with headers
import yfinance as yf
session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
})

def load_watchlists():
    print("🔧 load_watchlists() function was called!")
    global watchlists, target_hit_logged
    watchlists.clear()
    target_hit_logged.clear()

    try:
        if USE_GOOGLE_SHEETS:
            if not GOOGLE_SHEET_URL:
                raise ValueError("Google Sheets URL not set")
            
            csv_url = GOOGLE_SHEET_URL.replace("/edit?usp=sharing", "/export?format=csv")
            response = requests.get(csv_url)
            response.raise_for_status()

            df = pd.read_csv(
                StringIO(response.text),
                on_bad_lines='skip',
                engine='python',
                skipinitialspace=True,
                encoding='utf-8'
            )
            sheets = {"Watchlist 1": df}

        else:
            if not os.path.exists(DATA_FILE):
                raise FileNotFoundError(f"{DATA_FILE} not found")
            xls = pd.ExcelFile(DATA_FILE)
            sheets = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}

        for sheet_name, df in sheets.items():
            print(f"🔍 Raw columns in sheet '{sheet_name}': {list(df.columns)}")
            df.columns = df.columns.str.strip().str.lower()
            if 'scrip name' not in df or 'target price' not in df:
                print(f"⚠️ Sheet '{sheet_name}' missing 'Scrip Name' or 'Target Price' column")
                continue

            stocks = []
            target_hit_logged[sheet_name] = {}
            for _, row in df.iterrows():
                name = str(row['scrip name']).strip()
                if not name or name.lower() in ('nan', ''):
                    continue
                try:
                    target = float(row['target price'])
                except (ValueError, TypeError):
                    continue

                yf_symbol = name if name.endswith(('.NS', '.BO')) else f"{name}.NS"
                stocks.append({
                    "Scrip Name": name,
                    "Target Price": target,
                    "Current Price": 0.0,
                    "Status": "Loading...",
                    "yf_symbol": yf_symbol
                })
                target_hit_logged[sheet_name][name] = False

            if stocks:
                watchlists[sheet_name] = stocks
                print(f"✅ Loaded {len(stocks)} stocks from sheet '{sheet_name}'")
            else:
                print(f"⚠️ No valid stocks found in sheet '{sheet_name}'")

    except Exception as e:
        print(f"Error loading data: {e}")

def fetch_stock_price(symbol: str) -> float:
    print(f"🔍 Fetching price for: {symbol}")
    
    # Method 1: Try direct Yahoo Finance API with custom headers
    try:
        # Alternative: Direct Yahoo Finance API call
        base_symbol = symbol.replace('.NS', '').replace('.BO', '')
        
        # Try both NSE (.NS) and BSE (.BO) formats
        for suffix in ['.NS', '.BO']:
            test_symbol = f"{base_symbol}{suffix}"
            
            try:
                # Use custom session with headers
                ticker = yf.Ticker(test_symbol, session=session)
                
                # Try multiple methods to get price
                # Method 1: info (most reliable for Indian stocks)
                info = ticker.info
                if info and 'currentPrice' in info and info['currentPrice']:
                    price = float(info['currentPrice'])
                    print(f"✅ {test_symbol} → ₹{price:.2f} (info.currentPrice)")
                    return round(price, 2)
                
                if info and 'regularMarketPrice' in info and info['regularMarketPrice']:
                    price = float(info['regularMarketPrice'])
                    print(f"✅ {test_symbol} → ₹{price:.2f} (info.regularMarketPrice)")
                    return round(price, 2)
                
                # Method 2: history fallback
                hist = ticker.history(period="1d", interval="1m")
                if not hist.empty:
                    price = float(hist['Close'].iloc[-1])
                    print(f"✅ {test_symbol} → ₹{price:.2f} (history)")
                    return round(price, 2)
                    
            except Exception as e:
                print(f"⚠️ Failed {test_symbol}: {e}")
                continue
        
        # Method 2: Try alternative data source (NSE API)
        return fetch_from_nse_api(base_symbol)
        
    except Exception as e:
        print(f"⚠️ Error fetching {symbol}: {e}")
        return 0.0

def fetch_from_nse_api(symbol: str) -> float:
    """Alternative method using NSE API"""
    try:
        # NSE API endpoint
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
        
        # Try NSE quote API
        url = f"https://www.nseindia.com/api/quote-equity?symbol={symbol}"
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if 'priceInfo' in data and 'lastPrice' in data['priceInfo']:
                price = float(data['priceInfo']['lastPrice'])
                print(f"✅ {symbol} → ₹{price:.2f} (NSE API)")
                return round(price, 2)
                
    except Exception as e:
        print(f"⚠️ NSE API failed for {symbol}: {e}")
    
    # Method 3: Try Yahoo Finance with different approach
    try:
        # Fallback: Use requests directly to Yahoo Finance
        yahoo_url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}.NS"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        response = requests.get(yahoo_url, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if 'chart' in data and 'result' in data['chart'] and data['chart']['result']:
                result = data['chart']['result'][0]
                if 'meta' in result and 'regularMarketPrice' in result['meta']:
                    price = float(result['meta']['regularMarketPrice'])
                    print(f"✅ {symbol} → ₹{price:.2f} (Yahoo Direct)")
                    return round(price, 2)
    except Exception as e:
        print(f"⚠️ Yahoo Direct API failed for {symbol}: {e}")
    
    print(f"❌ All methods failed for {symbol}")
    return 0.0

def check_and_log_target_hit(sheet_name, scrip, target, current):
    if current >= target:
        if not target_hit_logged[sheet_name][scrip]:
            log_target_hit(sheet_name, scrip, target, current)
            target_hit_logged[sheet_name][scrip] = True
        return "🎯 Target Hit!"
    return "Below Target"

def log_target_hit(sheet_name, scrip_name, target_price, current_price):
    try:
        wb = load_workbook(LOG_FILE) if os.path.exists(LOG_FILE) else Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(['Scrip Name', 'Target Price', 'Hit Price', 'Date', 'Time', 'Timestamp'])
        else:
            ws = wb[sheet_name]

        now = datetime.now()
        ws.append([
            scrip_name,
            target_price,
            current_price,
            now.strftime('%Y-%m-%d'),
            now.strftime('%H:%M:%S'),
            now.strftime('%Y-%m-%d %H:%M:%S')
        ])
        wb.save(LOG_FILE)
        print(f"Logged: {scrip_name} hit target in {sheet_name}")
    except Exception as e:
        print(f"Log error: {e}")

def background_monitor():
    """Run every minute during market hours"""
    while monitoring_active:
        now = datetime.now()
        minute = now.minute
        weekday = now.weekday()
        current_time = now.time()
        market_start = datetime.strptime("09:15", "%H:%M").time()
        market_end = datetime.strptime("15:30", "%H:%M").time()

        should_run = (
            weekday < 5 and
            market_start <= current_time <= market_end and
            minute in [1, 16, 31, 46]
        )

        if should_run:
            print(f"[{now.strftime('%H:%M:%S')}] Auto-fetching all stocks...")
            for sheet_name, stocks in watchlists.items():
                for stock in stocks:
                    # Add small delay between requests to avoid rate limiting
                    time.sleep(0.5)
                    price = fetch_stock_price(stock["yf_symbol"])
                    stock["Current Price"] = price
                    status = check_and_log_target_hit(
                        sheet_name,
                        stock["Scrip Name"],
                        stock["Target Price"],
                        price
                    )
                    stock["Status"] = status
            print("Auto-fetch complete.")

        time_module.sleep(60)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/data")
def get_data():
    return jsonify(watchlists)

@app.route("/refresh")
def refresh():
    sheet = request.args.get("sheet")
    for sheet_name, stocks in watchlists.items():
        if sheet and sheet_name != sheet:
            continue
        for stock in stocks:
            # Add delay between requests to avoid rate limiting
            time.sleep(0.5)
            price = fetch_stock_price(stock["yf_symbol"])
            stock["Current Price"] = price
            status = check_and_log_target_hit(
                sheet_name,
                stock["Scrip Name"],
                stock["Target Price"],
                price
            )
            stock["Status"] = status
    return "", 204

@app.route("/reload")
def reload():
    load_watchlists()
    return "", 204

@app.route("/log")
def view_log():
    if os.path.exists(LOG_FILE):
        return redirect(f"/static/{LOG_FILE}")
    return "No log file yet.", 200

@app.route("/static/<path:filename>")
def static_file(filename):
    return send_from_directory(".", filename)

if __name__ == "__main__":
    if not os.path.exists(DATA_FILE):
        print(f"📁 {DATA_FILE} not found. Creating a sample file...")
        sample_df = pd.DataFrame({
            "Scrip Name": ["RELIANCE", "TATASTEEL", "INFY"],
            "Target Price": [3000.0, 180.0, 1500.0]
        })
        sample_df.to_excel(DATA_FILE, sheet_name="Watchlist 1", index=False)
        print("✅ Sample file created.")

    print("🔄 Loading watchlists...")
    load_watchlists()

    thread = threading.Thread(target=background_monitor, daemon=True)
    thread.start()

    port = int(os.environ.get("PORT", 5000))
    print(f"🌍 Open http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=True)
